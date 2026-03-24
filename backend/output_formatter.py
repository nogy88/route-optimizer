# ============================================================
#  output_formatter.py  –  Build API responses & Excel export
#  v6: Rural-specific styling removed (all routes are urban now).
#      "has_rural" / "is_rural" fields default to False safely.
# ============================================================

import io
import math
from datetime import datetime
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config


def _fmt_time(seconds) -> str:
    """Convert seconds → HH:MM (handles int, float, numpy scalars)."""
    s = int(seconds)
    return f"{s // 3600:02d}:{(s % 3600) // 60:02d}"


def _b(val) -> bool:
    return bool(val)


def _f(val, d=2) -> float:
    return round(float(val), d)


def _i(val) -> int:
    return int(val)


def _cost(route: Dict) -> Dict:
    v        = route["vehicle"]
    dist_km  = route["total_dist_m"] / 1000
    fuel     = dist_km * v["fuel_cost_km"]
    fixed    = v.get("vehicle_cost", 0)
    labor    = v.get("labor_cost", 0)
    return {"fuel": fuel, "fixed": fixed, "labor": labor,
            "total": fuel + fixed + labor}


def _depot_distances(store: Dict, dist_df) -> Dict:
    """Return distance (km) from store to each depot."""
    out = {}
    nid = store["node_id"]
    for dc_name in config.DEPOTS:
        try:
            row = dist_df.loc[dc_name, nid] if dc_name in dist_df.index else None
            if row is None:
                from data_loader import _norm_id
                n = _norm_id(dc_name)
                row = dist_df.loc[n, nid] if n in dist_df.index else None
            out[dc_name] = round(float(row) / 1000, 1) if row is not None else None
        except Exception:
            out[dc_name] = None
    return out


# ════════════════════════════════════════════════════════════
#  Summary builders
# ════════════════════════════════════════════════════════════

def build_route_summary(solver_result: Dict) -> List[Dict]:
    summaries = []
    for fleet, fr in solver_result.items():
        sched = config.FLEET_SCHEDULE.get(fleet, {})
        for route in fr["routes"]:
            c       = _cost(route)
            has_nd  = _b(any(s.get("is_next_day") for s in route["stops"]))
            cap_kg  = _f(route["cap_kg"])
            cap_m3  = _f(route["cap_m3"], 3)
            load_kg = _f(route["load_kg"])
            load_m3 = _f(route["load_m3"], 3)
            dist_m  = _f(route["total_dist_m"])
            dur_s   = _f(route["total_dur_s"])

            # Departure time: start_offset_s (shift-relative) → real wall-clock.
            # Trip 1: offset=0 → fleet start hour.  Trip N+1: real later departure.
            start_h       = sched.get("start_hour", 0)
            start_off_s   = route.get("start_offset_s", 0)
            depart_wall   = start_off_s + start_h * 3600
            departs_label = _fmt_time(depart_wall % 86400)

            # Return time (shift-relative → wall-clock)
            ret_s     = route.get("return_time_s", 0)
            ret_wall  = ret_s + start_h * 3600
            ret_label = _fmt_time(ret_wall % 86400)

            summaries.append({
                "fleet"        : str(fleet),
                "truck_id"     : str(route["truck_id"]),
                "trip_number"  : _i(route["trip_number"]),
                "route_type"   : "🏙 Urban",  # rural removed in v6
                "stops"        : _i(len(route["stops"])),
                "distance_km"  : round(dist_m / 1000, 2),
                "duration_min" : round(dur_s / 60, 1),
                "load_kg"      : load_kg,
                "cap_kg"       : cap_kg,
                "util_kg_pct"  : round(load_kg / cap_kg * 100, 1) if cap_kg else 0.0,
                "load_m3"      : load_m3,
                "cap_m3"       : cap_m3,
                "util_m3_pct"  : round(load_m3 / cap_m3 * 100, 1) if cap_m3 else 0.0,
                "cost_fuel"    : round(_f(c["fuel"]),  0),
                "cost_fixed"   : round(_f(c["fixed"]), 0),
                "cost_labor"   : round(_f(c["labor"]), 0),
                "cost_total"   : round(_f(c["total"]), 0),
                "departs_at"   : departs_label,
                "returns_at"   : ret_label,
                "is_overnight" : has_nd,
                # Man-hours consumed by this trip: depot-departure → depot-return.
                "man_hours"    : round(
                    max(0.0, route.get("return_time_s", 0)
                            - route.get("start_offset_s", 0)) / 3600.0, 2
                ),
            })
    return summaries


def build_stop_details(solver_result: Dict) -> List[Dict]:
    stops_out = []
    for fleet, fr in solver_result.items():
        for route in fr["routes"]:
            for order, stop in enumerate(route["stops"], 1):
                s       = stop["store"]
                arr_s   = _i(stop["arrival_s"])
                dep_s   = _i(stop["depart_s"])
                day_num = 1 + (arr_s // 86400)

                stops_out.append({
                    "fleet"       : str(fleet),
                    "truck_id"    : str(route["truck_id"]),
                    "trip_number" : _i(route["trip_number"]),
                    "stop_order"  : _i(order),
                    "store_id"    : str(s["store_id"]),
                    "eng_name"    : str(s.get("eng_name") or ""),
                    "mn_name"     : str(s.get("mn_name")  or ""),
                    "address"     : str(s.get("address")  or ""),
                    "detail_addr" : str(s.get("detail_addr") or ""),
                    "lat"         : _f(s["lat"], 6),
                    "lon"         : _f(s["lon"], 6),
                    "arrival"     : _fmt_time(arr_s % 86400),
                    "departure"   : _fmt_time(dep_s % 86400),
                    "delivery_day": f"Day {day_num}" if day_num > 1 else "Same day",
                    "is_rural"    : False,   # always False in v6
                    "demand_kg"   : _f(stop["demand_kg"]),
                    "demand_m3"   : _f(stop["demand_m3"], 3),
                })
    return stops_out


def build_unserved(solver_result: Dict, dist_df=None) -> List[Dict]:
    unserved_out = []
    seen = set()
    for fleet, fr in solver_result.items():
        for item in fr["unserved"]:
            s   = item["store"]
            key = (s["store_id"], fleet)
            if key in seen:
                continue
            seen.add(key)
            row = {
                "fleet"     : fleet,
                "store_id"  : s["store_id"],
                "eng_name"  : s.get("eng_name", ""),
                "mn_name"   : s.get("mn_name",  ""),
                "address"   : s.get("address",  ""),
                "lat"       : _f(s["lat"], 6),
                "lon"       : _f(s["lon"], 6),
                "demand_kg" : round(s["dry_kg"]  if fleet == "DRY" else s["cold_kg"],  2),
                "demand_m3" : round(s["dry_cbm"] if fleet == "DRY" else s["cold_cbm"], 3),
                "reason"    : item["reason"],
            }
            if dist_df is not None:
                dists = _depot_distances(s, dist_df)
                for dc, d in dists.items():
                    row[f"dist_from_{dc.replace(' ','_')}_km"] = d
            unserved_out.append(row)
    return unserved_out


def build_map_data(solver_result: Dict, route_geometries: Dict) -> List[Dict]:
    """Build per-route map data for frontend rendering."""
    COLORS = [
        "#5B7CFA","#22D3EE","#34D399","#A78BFA","#F472B6",
        "#38BDF8","#4ADE80","#818CF8","#FB7185","#2DD4BF",
        "#F97316","#EAB308","#84CC16","#DC2626","#D97706",
    ]

    routes_map = []
    color_idx  = {"DRY": 0, "COLD": 8}   # offset cold fleet colors

    for fleet, fr in solver_result.items():
        sched = config.FLEET_SCHEDULE.get(fleet, {})
        for route in fr["routes"]:
            rid = route["virtual_id"]
            idx = color_idx[fleet] % len(COLORS)
            color_idx[fleet] += 1
            color = COLORS[idx]

            depot_coords = (
                config.DEPOTS["Dry DC"] if fleet == "DRY"
                else config.DEPOTS["Cold DC"]
            )
            waypoints = [[depot_coords["lat"], depot_coords["lon"]]]

            stop_markers = []
            for i, stop in enumerate(route["stops"]):
                waypoints.append([stop["lat"], stop["lon"]])
                s         = stop["store"]
                arr_s     = stop["arrival_s"]
                day_num   = 1 + (int(arr_s) // 86400)
                day_label = f"Day {day_num}" if day_num > 1 else ""

                stop_markers.append({
                    "lat"         : _f(stop["lat"], 6),
                    "lon"         : _f(stop["lon"], 6),
                    "order"       : _i(i + 1),
                    "store_id"    : str(s["store_id"]),
                    "name"        : str(s.get("eng_name") or ""),
                    "mn_name"     : str(s.get("mn_name")  or ""),
                    "arrival"     : _fmt_time(_i(arr_s) % 86400),
                    "day_label"   : day_label,
                    "is_rural"    : False,
                    "is_next_day" : _b(stop.get("is_next_day", False)),
                    "demand_kg"   : _f(stop["demand_kg"]),
                    "demand_m3"   : _f(stop["demand_m3"], 3),
                })

            waypoints.append([depot_coords["lat"], depot_coords["lon"]])

            geometry = route_geometries.get(rid)
            polyline  = [[pt[1], pt[0]] for pt in geometry] if geometry else waypoints

            # Return-time label for tooltip
            ret_s    = route.get("return_time_s", 0)
            start_h  = sched.get("start_hour", 0)
            ret_wall = ret_s + start_h * 3600
            ret_lbl  = _fmt_time(ret_wall % 86400)

            routes_map.append({
                "route_id"    : rid,
                "fleet"       : fleet,
                "truck_id"    : route["truck_id"],
                "trip_number" : route["trip_number"],
                "is_rural"    : False,
                "color"       : color,
                "line_style"  : "solid",
                "stops"       : stop_markers,
                "polyline"    : polyline,
                "depot_lat"   : depot_coords["lat"],
                "depot_lon"   : depot_coords["lon"],
                "sched_info"  : (
                    f"Departs {sched.get('start_hour', 0):02d}:00 · "
                    f"Returns {ret_lbl}"
                ),
                "summary"     : {
                    "distance_km" : round(route["total_dist_m"] / 1000, 1),
                    "duration_min": round(route["total_dur_s"] / 60, 1),
                    "load_kg"     : round(route["load_kg"], 1),
                    "load_m3"     : round(route["load_m3"], 3),
                    "return_at"   : ret_lbl,
                    "is_overnight": _b(any(
                        s.get("is_next_day") for s in route["stops"]
                    )),
                }
            })
    return routes_map


# ════════════════════════════════════════════════════════════
#  Excel export
# ════════════════════════════════════════════════════════════

def _header_style():
    return {
        "font" : Font(bold=True, color="FFFFFF"),
        "fill" : PatternFill("solid", fgColor="1F3864"),
        "align": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }

def _apply_header(ws, headers):
    hs = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hs["font"]
        cell.fill      = hs["fill"]
        cell.alignment = hs["align"]
    ws.row_dimensions[1].height = 35

def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

def _alt_fill(row, col_count, ws, i):
    if i % 2 == 0:
        fill = PatternFill("solid", fgColor="EEF2FF")
        for c in range(1, col_count + 1):
            ws.cell(row=row, column=c).fill = fill


def export_to_excel(
    route_summary: List[Dict],
    stop_details : List[Dict],
    unserved     : List[Dict],
) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Route Summary ───────────────────────────────
    ws1 = wb.active
    ws1.title = "Route Summary"
    h1 = [
        "Fleet","Truck","Trip","Departs","Returns","Stops",
        "Dist (km)","Duration (min)",
        "Load kg","Cap kg","Util kg %",
        "Load m³","Cap m³","Util m³ %",
        "Fuel Cost ₮","Fixed Cost ₮","Labor Cost ₮","Total Cost ₮"
    ]
    _apply_header(ws1, h1)

    for i, r in enumerate(route_summary, 1):
        row = [
            r["fleet"], r["truck_id"], r["trip_number"],
            r.get("departs_at", ""), r.get("returns_at", ""),
            r["stops"], r["distance_km"], r["duration_min"],
            r["load_kg"], r["cap_kg"], r["util_kg_pct"],
            r["load_m3"], r["cap_m3"], r["util_m3_pct"],
            r["cost_fuel"], r["cost_fixed"], r["cost_labor"], r["cost_total"],
        ]
        for c, v in enumerate(row, 1):
            ws1.cell(row=i + 1, column=c, value=v)
        _alt_fill(i + 1, len(h1), ws1, i)
    _auto_width(ws1)

    # ── Sheet 2: Stop Details ────────────────────────────────
    ws2 = wb.create_sheet("Stop Details")
    h2 = [
        "Fleet","Truck","Trip","#",
        "Store ID","Eng Name","MN Name","Address","Detail Address",
        "Lat","Lon","Arrival","Departure","Delivery Day",
        "Demand kg","Demand m³"
    ]
    _apply_header(ws2, h2)

    for i, r in enumerate(stop_details, 1):
        row = [
            r["fleet"], r["truck_id"], r["trip_number"], r["stop_order"],
            r["store_id"], r["eng_name"], r["mn_name"],
            r["address"], r.get("detail_addr", ""),
            r["lat"], r["lon"],
            r["arrival"], r["departure"],
            r.get("delivery_day", "Same day"),
            r["demand_kg"], r["demand_m3"],
        ]
        for c, v in enumerate(row, 1):
            ws2.cell(row=i + 1, column=c, value=v)
        _alt_fill(i + 1, len(h2), ws2, i)
    _auto_width(ws2)

    # ── Sheet 3: Unserved ────────────────────────────────────
    ws3 = wb.create_sheet("Unserved Stores")
    h3 = [
        "Fleet","Store ID","Eng Name","MN Name","Address","Lat","Lon",
        "Demand kg","Demand m³","→ Dry DC (km)","→ Cold DC (km)","Reason"
    ]
    _apply_header(ws3, h3)
    red = PatternFill("solid", fgColor="FFE0E0")
    for i, r in enumerate(unserved, 1):
        row = [
            r["fleet"], r["store_id"], r["eng_name"], r["mn_name"],
            r["address"], r["lat"], r["lon"],
            r["demand_kg"], r["demand_m3"],
            r.get("dist_from_Dry_DC_km"),
            r.get("dist_from_Cold_DC_km"),
            r["reason"],
        ]
        for c, v in enumerate(row, 1):
            ws3.cell(row=i + 1, column=c, value=v).fill = red
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()